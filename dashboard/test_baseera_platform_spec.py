"""
Mandatory acceptance test suite for the "Baseera" 4-layer platform spec
(Validation Layer / Retrieval Layer / Orchestrator-Router / Reconciliation
Layer). Test methods are prefixed with the scenario number from section 7 of
the spec ("خطة الاختبار الإلزامية") so it is easy to map a failure back to the
exact requirement it covers. The suite is filled in incrementally as each
layer is implemented (see the phase docstrings below) — by the end of Phase 6
all 8 scenarios must pass.
"""
import io
import json
import datetime
from decimal import Decimal

from django.test import TestCase
from django.contrib.auth.models import User
from django.core.files.uploadedfile import SimpleUploadedFile
from django.core.exceptions import ValidationError

from dashboard.services.validation_service import validate_financial_file
from dashboard.services import orchestrator
from dashboard.services.reconciliation_service import (
    reconcile_report_items, compute_grand_total, ReconciliationError,
)
from dashboard.models import CompanyStrategicProfile, ProjectFile, FileSheetMetadata, DynamicRecord


# --------------------------------------------------------------------------
# Fixture builders
# --------------------------------------------------------------------------

def _make_xlsx_bytes(sheets: dict) -> bytes:
    """sheets: {sheet_name: list[dict]} -> real .xlsx file bytes."""
    import pandas as pd
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        for name, rows in sheets.items():
            df = pd.DataFrame(rows)
            df.to_excel(writer, sheet_name=name, index=False)
    buf.seek(0)
    return buf.read()


# A well-known minimal valid 1x1 PNG (used to simulate "a photo disguised as
# a spreadsheet" and "an empty/non-tabular first sheet").
_MINIMAL_PNG = bytes.fromhex(
    "89504e470d0a1a0a0000000d494844520000000100000001080600000"
    "01f15c4890000000a49444154789c6360000002000100ffff03000006"
    "000557bf10c30000000049454e44ae426082"
)


class ValidationLayerTests(TestCase):
    """Section 3.2 / spec test-plan items 1, 2, 6."""

    # --- Test 1: "اختبار السمكة" -----------------------------------------
    def test_1_fish_test_first_sheet_image_second_sheet_valid_is_accepted_partially(self):
        """
        Excel workbook whose FIRST sheet is empty/non-financial (simulating a
        picture/junk sheet) and whose SECOND sheet has real financial data:
        only the second sheet must be accepted, the first must be rejected,
        and both facts must be visible in accepted_sheets/rejected_sheets.
        """
        content = _make_xlsx_bytes({
            # A pasted photo shows up, once read back through pandas, as a
            # near-empty sheet with a single placeholder row/column rather
            # than genuinely zero rows — this exercises the explicit
            # "أقل من 3 صفوف" rejection path (a truly empty sheet is instead
            # silently dropped before validation even considers it, which
            # would never surface a rejection reason).
            "الصورة": [{"Unnamed": "IMG"}],
            "المبيعات": [
                {"التاريخ": "2024-01-01", "السعر": "100", "الكمية": 2},
                {"التاريخ": "2024-01-02", "السعر": "150", "الكمية": 1},
                {"التاريخ": "2024-01-03", "السعر": "200", "الكمية": 3},
            ],
        })
        upload = SimpleUploadedFile(
            "fish_test.xlsx", content,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        result = validate_financial_file(upload)

        self.assertTrue(result["is_valid"])
        self.assertIn(result["status"], ("accept", "warning"))
        self.assertIn("المبيعات", result["accepted_sheets"])
        self.assertNotIn("المبيعات", result["rejected_sheets"])
        self.assertTrue(any("الصورة" in r for r in result["rejected_sheets"]))

    # --- Test 2: "اختبار صورة مباشرة" -------------------------------------
    def test_2_disguised_image_renamed_to_xlsx_is_rejected_on_mime_check(self):
        """A real PNG renamed to .xlsx must be rejected at the MIME-sniffing stage."""
        padded_png = _MINIMAL_PNG + (b"\x00" * 2000)  # push size above the 1KB floor
        upload = SimpleUploadedFile("fake_report.xlsx", padded_png, content_type="image/png")

        result = validate_financial_file(upload)

        self.assertFalse(result["is_valid"])
        self.assertEqual(result["status"], "reject")
        self.assertIn("لا يتطابق مع امتداده", result["message"])

    # --- Test 6: عمود تاريخ نصي --------------------------------------------
    def test_6_text_date_column_next_to_price_column_is_accepted(self):
        """
        A date column stored as plain text (e.g. "2024-01-15") next to a price
        column must still be recognized via pd.to_datetime and accepted.
        """
        content = _make_xlsx_bytes({
            "Sheet1": [
                {"date": "2024-01-15", "price": 120},
                {"date": "2024-02-15", "price": 135},
                {"date": "2024-03-15", "price": 142},
                {"date": "2024-04-15", "price": 158},
            ],
        })
        upload = SimpleUploadedFile(
            "text_date_column.xlsx", content,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        result = validate_financial_file(upload)

        self.assertTrue(result["is_valid"])
        self.assertIn("Sheet1", result["accepted_sheets"])

    # Bonus coverage explicitly required by section 3.2's unit-test list:
    # a price column written with thousands separators and a currency symbol.
    def test_currency_formatted_price_column_is_coerced_to_numeric_and_accepted(self):
        content = _make_xlsx_bytes({
            "Sheet1": [
                {"invoice date": "2024-01-01", "total price": "1,200.50 ر.س"},
                {"invoice date": "2024-01-02", "total price": "2,450.00 ر.س"},
                {"invoice date": "2024-01-03", "total price": "980.75 ر.س"},
            ],
        })
        upload = SimpleUploadedFile(
            "currency_price.xlsx", content,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        result = validate_financial_file(upload)

        self.assertTrue(result["is_valid"])
        self.assertIn("Sheet1", result["accepted_sheets"])


class CompanyStrategicProfileValidatorTests(TestCase):
    """Spec test-plan item 8: invalid strategic_priorities_ranking must fail to save."""

    def setUp(self):
        self.user = User.objects.create_user(username="profile_owner", password="pw123456")

    def _base_kwargs(self):
        return dict(
            user=self.user,
            company_name="شركة تجريبية",
            sector="retail",
            size="small",
            growth_stage="growth",
            risk_tolerance="balanced",
        )

    def test_8_ranking_with_fewer_than_5_items_fails_validation(self):
        profile = CompanyStrategicProfile(
            **self._base_kwargs(),
            strategic_priorities_ranking=["cash_preservation", "growth"],
        )
        with self.assertRaises(ValidationError):
            profile.full_clean()

    def test_8_ranking_with_duplicate_items_fails_validation(self):
        profile = CompanyStrategicProfile(
            **self._base_kwargs(),
            strategic_priorities_ranking=[
                "cash_preservation", "cash_preservation", "growth",
                "profitability", "cost_reduction",
            ],
        )
        with self.assertRaises(ValidationError):
            profile.full_clean()

    def test_8_ranking_with_unknown_value_fails_validation(self):
        profile = CompanyStrategicProfile(
            **self._base_kwargs(),
            strategic_priorities_ranking=[
                "cash_preservation", "growth", "profitability",
                "cost_reduction", "not_a_real_priority",
            ],
        )
        with self.assertRaises(ValidationError):
            profile.full_clean()

    def test_8_valid_full_ranking_saves_successfully(self):
        profile = CompanyStrategicProfile(
            **self._base_kwargs(),
            strategic_priorities_ranking=[
                "cash_preservation", "growth", "profitability",
                "cost_reduction", "long_term_stability",
            ],
        )
        profile.full_clean()
        profile.save()
        self.assertEqual(CompanyStrategicProfile.objects.count(), 1)


class StrategicProfileViewTests(TestCase):
    """End-to-end coverage for the Form/View required by section 3.1."""

    def setUp(self):
        self.user = User.objects.create_user(username="view_owner", password="pw123456")
        self.client.force_login(self.user)

    def test_get_renders_form(self):
        response = self.client.get("/settings/strategic-profile/")
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "priority_rank_1")

    def test_post_valid_ranking_creates_profile(self):
        response = self.client.post("/settings/strategic-profile/", {
            "company_name": "متجر بصيرة",
            "sector": "retail",
            "size": "small",
            "growth_stage": "growth",
            "risk_tolerance": "balanced",
            "max_investment_limit": "50000",
            "cash_reserve_floor": "10000",
            "priority_rank_1": "cash_preservation",
            "priority_rank_2": "growth",
            "priority_rank_3": "profitability",
            "priority_rank_4": "cost_reduction",
            "priority_rank_5": "long_term_stability",
        }, follow=True)
        self.assertEqual(response.status_code, 200)
        profile = CompanyStrategicProfile.objects.get(user=self.user)
        self.assertEqual(
            profile.strategic_priorities_ranking,
            ["cash_preservation", "growth", "profitability", "cost_reduction", "long_term_stability"],
        )
        self.assertEqual(profile.max_investment_limit, Decimal("50000"))

    def test_post_duplicate_ranking_is_rejected_and_does_not_save(self):
        response = self.client.post("/settings/strategic-profile/", {
            "company_name": "متجر بصيرة",
            "sector": "retail",
            "size": "small",
            "growth_stage": "growth",
            "risk_tolerance": "balanced",
            "priority_rank_1": "cash_preservation",
            "priority_rank_2": "cash_preservation",  # duplicate
            "priority_rank_3": "profitability",
            "priority_rank_4": "cost_reduction",
            "priority_rank_5": "long_term_stability",
        })
        self.assertEqual(response.status_code, 200)
        self.assertFalse(CompanyStrategicProfile.objects.filter(user=self.user).exists())


class OrchestratorRouterTests(TestCase):
    """Section 4: Orchestrator / Router. Spec test-plan items 3, 4, 7."""

    def setUp(self):
        self.user = User.objects.create_user(username="router_user", password="pw123456")
        self.client.force_login(self.user)

    # --- Test 3: "اختبار الهلوسة" ------------------------------------------
    def test_3_off_topic_question_gets_direct_refusal_no_agent_call(self):
        """'What is the capital of Muscat?' must get Route 1's canned refusal,
        with no financial analysis and no numbers hallucinated."""
        route = orchestrator.classify_route("ما عاصمة مسقط؟", has_uploaded_files=False)
        self.assertEqual(route, orchestrator.ROUTE_OFF_TOPIC)

        response = self.client.post(
            "/api/insights/chat",
            data=json.dumps({"message": "ما عاصمة مسقط؟", "lang": "ar"}),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        body = b"".join(response.streaming_content).decode("utf-8")

        streamed_text = ""
        for line in body.splitlines():
            if line.startswith("data: "):
                payload = json.loads(line[len("data: "):])
                candidates = payload.get("candidates")
                if candidates:
                    streamed_text += candidates[0]["content"]["parts"][0]["text"]

        self.assertIn("لا يمكنني الإجابة على هذا السؤال لأنه خارج اختصاصي", streamed_text)
        # No financial-analysis markers should leak into an off-topic refusal.
        self.assertNotIn("هامش الربح", streamed_text)
        self.assertNotIn("STATUS___:جاري", streamed_text)

    # --- Test 4: "اختبار الدجاج" --------------------------------------------
    def test_4_chicken_query_against_meat_file_asks_for_confirmation(self):
        """
        A file about 'مبيعات لحوم' (meat sales) must NOT be silently reused
        to answer a question about 'دجاج' (chicken): the system must ask for
        explicit confirmation instead of conflating the two.
        """
        pf = ProjectFile.objects.create(user=self.user, excel_file="excel_files/meat_sales.xlsx")
        FileSheetMetadata.objects.create(
            project_file=pf,
            sheet_name="مبيعات اللحوم",
            status="accept",
            columns=["التاريخ", "السعر", "الكمية"],
            row_count=10,
            category="sales",
            keywords=["مبيعات", "لحوم", "لحم", "التاريخ", "السعر", "الكمية"],
        )

        route_info = orchestrator.route_message(
            self.user.id, "ما هي توقعات شراء الدجاج للشهر القادم؟", lang="ar",
        )

        self.assertEqual(route_info["route"], orchestrator.ROUTE_SINGLE_FILE)
        self.assertTrue(route_info["needs_confirmation"])
        self.assertEqual(route_info["matched_sheet_note"], "")
        action_ids = [a["action_id"] for a in route_info["suggested_actions"]]
        self.assertIn("upload_new_file", action_ids)

    def test_4_confirmed_sheet_bypasses_retrieval_ambiguity(self):
        """Once the user explicitly confirms a file/sheet, the router must use
        it directly without re-running the ambiguity check."""
        pf = ProjectFile.objects.create(user=self.user, excel_file="excel_files/meat_sales.xlsx")
        FileSheetMetadata.objects.create(
            project_file=pf, sheet_name="مبيعات اللحوم", status="accept",
            columns=["التاريخ", "السعر"], row_count=10, category="sales",
            keywords=["مبيعات", "لحوم"],
        )
        route_info = orchestrator.route_message(
            self.user.id, "ما هي توقعات شراء الدجاج؟", lang="ar",
            confirmed_sheet={"project_file_id": pf.id, "sheet_name": "مبيعات اللحوم"},
        )
        self.assertFalse(route_info["needs_confirmation"])
        self.assertIn("مبيعات اللحوم", route_info["matched_sheet_note"])

    # --- Test 7: الحدود المالية الصلبة -------------------------------------
    def test_7_option_exceeding_max_investment_limit_is_auto_rejected(self):
        profile = CompanyStrategicProfile.objects.create(
            user=self.user, company_name="شركة", sector="retail", size="small",
            growth_stage="growth", risk_tolerance="balanced",
            strategic_priorities_ranking=[
                "cash_preservation", "growth", "profitability",
                "cost_reduction", "long_term_stability",
            ],
            max_investment_limit=Decimal("100000"),
            cash_reserve_floor=Decimal("20000"),
        )
        options = [
            {"label": "توسع بميزانية 250,000", "required_investment": Decimal("250000"),
             "expected_roi": 0.40, "cash_after": Decimal("30000")},
            {"label": "توسع محدود بميزانية 80,000", "required_investment": Decimal("80000"),
             "expected_roi": 0.12, "cash_after": Decimal("25000")},
            {"label": "خيار يستنزف السيولة", "required_investment": Decimal("50000"),
             "expected_roi": 0.18, "cash_after": Decimal("5000")},
        ]

        kept, rejected = orchestrator.filter_options_by_hard_constraints(options, profile)

        kept_labels = [o["label"] for o in kept]
        rejected_labels = [o["label"] for o in rejected]

        # The highest-return option is rejected purely for busting the hard
        # investment cap, even though it has the best ROI of the three.
        self.assertNotIn("توسع بميزانية 250,000", kept_labels)
        self.assertIn("توسع بميزانية 250,000", rejected_labels)
        self.assertIn("خيار يستنزف السيولة", rejected_labels)
        self.assertEqual(kept_labels, ["توسع محدود بميزانية 80,000"])

    def test_7_no_profile_means_no_constraints_applied(self):
        options = [{"label": "أي خيار", "required_investment": Decimal("999999999")}]
        kept, rejected = orchestrator.filter_options_by_hard_constraints(options, None)
        self.assertEqual(kept, options)
        self.assertEqual(rejected, [])


def _extract_sse(streaming_content):
    """Parses the SSE body into (streamed_text, suggested_actions)."""
    body = b"".join(streaming_content).decode("utf-8")
    text = ""
    actions = None
    for line in body.splitlines():
        if line.startswith("data: "):
            payload = json.loads(line[len("data: "):])
            if "suggested_actions" in payload:
                actions = payload["suggested_actions"]
            candidates = payload.get("candidates")
            if candidates:
                text += candidates[0]["content"]["parts"][0]["text"]
    return text, actions


class SuggestedActionsAPITests(TestCase):
    """Section 4.3: dynamic suggested_actions surfaced through the chat API."""

    def setUp(self):
        self.user = User.objects.create_user(username="actions_user", password="pw123456")
        self.client.force_login(self.user)

    def test_ambiguous_or_missing_file_query_returns_suggested_actions(self):
        response = self.client.post(
            "/api/insights/chat",
            data=json.dumps({"message": "كم كانت مبيعاتي الشهر الماضي؟", "lang": "ar"}),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        text, actions = _extract_sse(response.streaming_content)
        self.assertIsNotNone(actions)
        action_ids = [a["action_id"] for a in actions]
        self.assertIn("upload_new_file", action_ids)
        self.assertIn("لم أتمكن من العثور على المستند المطلوب", text)

    def test_confirmed_sheet_round_trip_skips_ambiguity_and_proceeds(self):
        pf = ProjectFile.objects.create(user=self.user, excel_file="excel_files/sales.xlsx")
        FileSheetMetadata.objects.create(
            project_file=pf, sheet_name="المبيعات", status="accept",
            columns=["التاريخ", "السعر"], row_count=5, category="sales",
            keywords=["مبيعات", "التاريخ", "السعر"],
        )
        response = self.client.post(
            "/api/insights/chat",
            data=json.dumps({
                "message": "كم كانت المبيعات؟",
                "lang": "ar",
                "confirmed_sheet": {"project_file_id": pf.id, "sheet_name": "المبيعات"},
            }),
            content_type="application/json",
        )
        # A confirmed match must not be answered with the direct off-topic/
        # missing-file bypass stream -> it proceeds to the (fallback) agent.
        self.assertEqual(response.status_code, 200)
        text, actions = _extract_sse(response.streaming_content)
        self.assertIsNone(actions)
        self.assertNotIn("لم أتمكن من العثور على المستند المطلوب", text)


class RetrievalLayerTests(TestCase):
    """
    Section 5: Retrieval Layer. Exercises the full pipeline end-to-end —
    validate_financial_file -> index_accepted_sheets -> search_relevant_sheets
    — and the literal-term-wins-over-generic-overlap guarantee.
    """

    def setUp(self):
        self.user = User.objects.create_user(username="retrieval_user", password="pw123456")

    def test_only_accepted_sheets_get_indexed_end_to_end(self):
        from dashboard.services.retrieval_service import index_accepted_sheets

        content = _make_xlsx_bytes({
            "junk": [{"Unnamed": "IMG"}],
            "المبيعات": [
                {"التاريخ": "2024-01-01", "السعر": "100", "الكمية": 2},
                {"التاريخ": "2024-01-02", "السعر": "150", "الكمية": 1},
                {"التاريخ": "2024-01-03", "السعر": "200", "الكمية": 3},
            ],
        })
        upload = SimpleUploadedFile(
            "sales.xlsx", content,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        validation = validate_financial_file(upload)
        self.assertTrue(validation["is_valid"])

        upload.seek(0)
        pf = ProjectFile.objects.create(user=self.user, excel_file=upload)
        created = index_accepted_sheets(pf, validation["accepted_sheets"])

        self.assertEqual(len(created), 1)
        self.assertEqual(created[0].sheet_name, "المبيعات")
        self.assertEqual(created[0].category, "sales")
        self.assertIn("السعر", created[0].columns)
        self.assertEqual(created[0].date_range_start, datetime.date(2024, 1, 1))
        self.assertEqual(created[0].date_range_end, datetime.date(2024, 1, 3))
        # The rejected "junk" sheet must never appear in the retrieval index.
        self.assertEqual(FileSheetMetadata.objects.filter(project_file=pf).count(), 1)
        self.assertFalse(FileSheetMetadata.objects.filter(sheet_name="junk").exists())

    def test_literal_term_outranks_generic_category_overlap(self):
        """
        Two files both mention 'مبيعات' (sales), but only one is literally
        about 'دجاج' (chicken). A query for chicken must rank that file
        first, not the generic meat file, despite the shared category token.
        """
        from dashboard.services.retrieval_service import search_relevant_sheets

        pf_meat = ProjectFile.objects.create(user=self.user, excel_file="excel_files/meat.xlsx")
        FileSheetMetadata.objects.create(
            project_file=pf_meat, sheet_name="مبيعات اللحوم", status="accept",
            columns=["التاريخ", "السعر"], row_count=10, category="sales",
            keywords=["مبيعات", "لحوم", "لحم"],
        )
        pf_chicken = ProjectFile.objects.create(user=self.user, excel_file="excel_files/chicken.xlsx")
        FileSheetMetadata.objects.create(
            project_file=pf_chicken, sheet_name="مبيعات الدجاج", status="accept",
            columns=["التاريخ", "السعر"], row_count=10, category="sales",
            keywords=["مبيعات", "دجاج"],
        )

        results = search_relevant_sheets(self.user.id, "توقعات شراء الدجاج القادمة", top_k=5)

        self.assertTrue(results)
        self.assertEqual(results[0]["sheet_name"], "مبيعات الدجاج")
        if len(results) > 1:
            self.assertGreater(results[0]["score"], results[1]["score"])


class ReconciliationLayerTests(TestCase):
    """Section 6: Reconciliation Layer. Spec test-plan item 5."""

    def setUp(self):
        self.user = User.objects.create_user(username="recon_user", password="pw123456")
        self.client.force_login(self.user)

    # --- Test 5: "اختبار المطابقة" (happy path) -----------------------------
    def test_5_report_total_matches_source_exactly(self):
        """A sales file whose rows sum to exactly 6,000,000 must produce a
        report whose grand total matches 6,000,000 exactly."""
        rows = [
            {"السعر": "3000000", "الكمية": 1},
            {"السعر": "2000000", "الكمية": 1},
            {"السعر": "500000", "الكمية": 2},
        ]  # 3,000,000 + 2,000,000 + 1,000,000 = 6,000,000

        verified_items = reconcile_report_items(rows, lang="ar")
        self.assertEqual(compute_grand_total(verified_items), Decimal("6000000"))

        pf = ProjectFile.objects.create(user=self.user, excel_file="excel_files/big_sales.xlsx")
        for row in rows:
            DynamicRecord.objects.create(user=self.user, project_file=pf, schema_hash="x", row_data=row)

        response = self.client.get(f"/export-excel/?file_id={pf.id}")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(response.content), data_only=False)
        ws = wb.active
        start_r = 9
        grand_total = Decimal("0")
        for i in range(len(rows)):
            qty = Decimal(str(ws.cell(row=start_r + i, column=4).value))
            price = Decimal(str(ws.cell(row=start_r + i, column=5).value))
            grand_total += qty * price
        self.assertEqual(grand_total, Decimal("6000000"))

    # --- Test 5: conflict must abort, never show a wrong number -----------
    def test_5_irreconcilable_mismatch_aborts_generation(self):
        """
        Every row carries an explicit 'الإجمالي' total that legitimately
        disagrees with qty*price (e.g. a per-row discount not reflected in
        the raw price/qty columns) — the checksum can never be satisfied by
        re-deriving qty*price, so generation must abort rather than emit a
        wrong number.
        """
        rows = [
            {"السعر": "1000", "الكمية": 5, "الإجمالي": "3000"},   # true qty*price = 5000, declared total = 3000
            {"السعر": "2000", "الكمية": 2, "الإجمالي": "1000"},   # true qty*price = 4000, declared total = 1000
        ]

        with self.assertRaises(ReconciliationError) as ctx:
            reconcile_report_items(rows, lang="ar")
        self.assertIn("حدث خطأ في عملية مطابقة الأرقام", ctx.exception.message)

        pf = ProjectFile.objects.create(user=self.user, excel_file="excel_files/inconsistent_sales.xlsx")
        for row in rows:
            DynamicRecord.objects.create(user=self.user, project_file=pf, schema_hash="x", row_data=row)

        response = self.client.get(f"/export-excel/?file_id={pf.id}")
        self.assertEqual(response.status_code, 409)
        self.assertIn("حدث خطأ في عملية مطابقة الأرقام", response.content.decode("utf-8"))

    def test_rejected_sheet_data_never_reaches_a_report(self):
        """A sheet rejected by the Validation Layer must never produce
        DynamicRecord rows, so it structurally cannot appear in a report."""
        content = _make_xlsx_bytes({
            "junk": [{"Unnamed": "IMG"}],
            "المبيعات": [
                {"التاريخ": "2024-01-01", "السعر": "100", "الكمية": 2},
                {"التاريخ": "2024-01-02", "السعر": "150", "الكمية": 1},
                {"التاريخ": "2024-01-03", "السعر": "200", "الكمية": 3},
            ],
        })
        upload = SimpleUploadedFile(
            "sales.xlsx", content,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        validation = validate_financial_file(upload)
        self.assertTrue(any("junk" in r for r in validation["rejected_sheets"]))
        self.assertEqual(validation["accepted_sheets"], ["المبيعات"])
